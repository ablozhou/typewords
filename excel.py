import openpyxl


class CasesData:
    """用于保存测试用例数据"""
    pass


class ReadExcel:
    def __init__(self, file_name):
        self.file_name = file_name
        self.dicts = {}

    def open(self):
        self.wb = openpyxl.load_workbook(self.file_name)
        self.sheets = self.wb.sheetnames

    def close(self):
        self.wb.close()

    def sheet_names(self):
        sheets={}
        self.open()
        for i in range(len(self.sheets)):
            sh = self.wb[self.sheets[i]]
            sheets[i]=sh.title
        # for all sheets
        sheets[99]="all"
        self.close()
        return sheets
        

    def read_data(self, sheet_id):
        """按行读取数据，最后返回一个字典"""
        sheets = []
        
        self.dicts={}
        self.open()
        if sheet_id >= 0 and sheet_id < len(self.sheets):
            sheets = [self.sheets[sheet_id]]
            print("practice %s"%sheets[0])
        else:
            sheet_id = 99
            print("practise all words")
            sheets = self.sheets            

        for i in range(len(sheets)):
            sh = self.wb[sheets[i]]
            rows = list(sh.rows)

            for row in rows:
                item = []
                for r in row:
                    #print(r.value)
                    if r.value:
                        item.append(r.value.strip())
                # print(item)

                self.dicts[item[0]] = item[1]
        self.close()

        return self.dicts

    def write_data(self, row, column, msg):
        self.open()
        self.sh.cell(row=row, column=column, value=msg)
        self.wb.save(self.file_name)
        self.close()


if __name__ == '__main__':
    # 需要读取excel时直接调用ReadExcel类
    words = ReadExcel('words.xlsx')
    d = words.read_data()
    print(d)

def testxlsx():
    dicts = {}
    with open('words.xlsx', 'rb') as f:
        wb = openpyxl.load_workbook(filename=f)

        # 获取workbook中所有的表格

        sheets = wb.sheetnames

        print(sheets)

        # 循环遍历所有sheet
        #for t in range(len(sheets)):
        for t in range(0, 1):

            sheet = wb[sheets[t]]
            print('\n\n第' + str(t + 1) + '个sheet: ' + sheet.title + '->>')

            #len_row代表表中有多少行数据，len_column代表excel表中有多少列数据
            len_row = sheet.max_row
            len_column = sheet.max_column
            for i in range(1, len_row + 1):
                dicts[sheet.cell(row=i,column=1).value] = \
                sheet.cell(row=i,column=2).value

        print(dicts)
